# -*- coding: utf-8 -*-
import openpyxl

def export_to_excel(review_entries, settings, file_path):
    """
    Exports review entries to an Excel file with only text information.
    """
    if not file_path:
        return False, "没有提供文件路径"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "审阅意见"
        
        ws.append([f"项目名称: {settings.get('project_name', '')}", f"制作人: {settings.get('producer', '')}", f"审阅人: {settings.get('reviewer', '')}"])
        ws.append([])
        
        # FIX: Simplified headers
        headers = ["镜头号", "时间戳", "完整意见", "简化意见", "关键词", "制作部门"]
        ws.append(headers)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25 # 镜头号
        ws.column_dimensions['B'].width = 15 # 时间戳
        ws.column_dimensions['C'].width = 60 # 完整意见
        ws.column_dimensions['D'].width = 60 # 简化意见
        ws.column_dimensions['E'].width = 30 # 关键词
        ws.column_dimensions['F'].width = 15 # 制作部门

        for entry in review_entries:
            # Wrap text for long reviews
            full_review_cell = ws.cell(row=ws.max_row + 1, column=3, value=entry.get("full_review", ""))
            full_review_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            
            simplified_review_cell = ws.cell(row=ws.max_row, column=4, value=entry.get("simplified_review", ""))
            simplified_review_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

            ws.cell(row=ws.max_row, column=1, value=entry.get("shot_number", ""))
            ws.cell(row=ws.max_row, column=2, value=entry.get("timestamp", ""))
            ws.cell(row=ws.max_row, column=5, value=", ".join(entry.get("keywords", [])))
            ws.cell(row=ws.max_row, column=6, value=entry.get("department", ""))

        wb.save(file_path)
        return True, file_path
    except Exception as e:
        return False, str(e)
